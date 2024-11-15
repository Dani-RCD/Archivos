from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Crear y escribir en un archivo de Excel
wb = Workbook()
ws = wb.active
ws.title = "Datos de Personas"
ws["A1"] = "Nombre"
ws["B1"] = "Edad"
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
ws["A1"].font = header_font
ws["B1"].font = header_font
ws["A1"].fill = header_fill
ws["B1"].fill = header_fill
ws["A1"].alignment = Alignment(horizontal="center")
ws["B1"].alignment = Alignment(horizontal="center")
data = [["Juan", 28], ["Ana", 22], ["Luis", 34], ["Marta", 45]]
for row in data:
    ws.append(row)
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 10
wb.save("archivoDRC.xlsx")
# Leer y mostrar el archivo de Excel
wb = load_workbook("archivoDRC.xlsx")
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
